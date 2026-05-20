import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_INPUT_FILE = "superstore_clean.csv"
OUTPUT_FILE = "RetailPulse_Report.xlsx"
REQUIRED_COLUMNS = {
    "Order ID",
    "Order Date",
    "Sales",
    "Profit",
    "Region",
    "Category",
    "Customer ID",
    "Customer Name",
    "Segment",
    "Product Name",
}

CURRENCY_FORMAT = '"$"#,##0.00'
PERCENT_FORMAT = "0.00%"
INTEGER_FORMAT = "#,##0"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
KPI_FILL = PatternFill("solid", fgColor="EAF3E2")
ALT_ROW_FILL = PatternFill("solid", fgColor="F7FBFF")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def resolve_input_path() -> Path:
    if len(sys.argv) > 1:
        return Path(sys.argv[1]).expanduser().resolve()
    return (Path.cwd() / DEFAULT_INPUT_FILE).resolve()


def ensure_required_columns(df: pd.DataFrame) -> None:
    missing_columns = REQUIRED_COLUMNS.difference(df.columns)
    if missing_columns:
        missing_list = ", ".join(sorted(missing_columns))
        raise ValueError(f"Missing required columns: {missing_list}")


def prepare_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    prepared = df.copy()
    prepared["Order Date"] = pd.to_datetime(prepared["Order Date"], errors="coerce")
    prepared = prepared.dropna(subset=["Order Date"]).copy()

    if "Year" not in prepared.columns:
        prepared["Year"] = prepared["Order Date"].dt.year

    if "Profit Margin %" not in prepared.columns:
        sales_nonzero = prepared["Sales"].replace(0, pd.NA)
        prepared["Profit Margin %"] = ((prepared["Profit"] / sales_nonzero) * 100).fillna(0)

    return prepared


def safe_margin(profit: pd.Series, sales: pd.Series) -> pd.Series:
    return (profit / sales.replace(0, pd.NA)).fillna(0)


def style_header_row(worksheet, row_number: int) -> None:
    for cell in worksheet[row_number]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_data_range(worksheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in worksheet.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    ):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")


def autofit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 28)


def create_yearly_summary(df: pd.DataFrame) -> pd.DataFrame:
    yearly = (
        df.groupby("Year", dropna=False)
        .agg(
            total_revenue=("Sales", "sum"),
            total_profit=("Profit", "sum"),
            total_orders=("Order ID", "nunique"),
        )
        .reset_index()
        .sort_values("Year")
    )
    yearly["avg_order_value"] = (
        yearly["total_revenue"] / yearly["total_orders"].replace(0, pd.NA)
    ).fillna(0)
    yearly["profit_margin"] = safe_margin(yearly["total_profit"], yearly["total_revenue"])
    yearly["yoy_growth"] = yearly["total_revenue"].pct_change().fillna(0)
    return yearly


def create_regional_pivot(df: pd.DataFrame) -> pd.DataFrame:
    pivot = pd.pivot_table(
        df,
        index="Region",
        columns="Category",
        values="Sales",
        aggfunc="sum",
        fill_value=0,
    )
    pivot["Total Sales"] = pivot.sum(axis=1)

    region_profit = df.groupby("Region", dropna=False)["Profit"].sum()
    pivot["Total Profit"] = region_profit.reindex(pivot.index).fillna(0)
    pivot["Profit Margin"] = safe_margin(pivot["Total Profit"], pivot["Total Sales"])

    total_row = pd.DataFrame([pivot.sum(numeric_only=True)], index=["Grand Total"])
    total_row["Profit Margin"] = (
        total_row["Total Profit"] / total_row["Total Sales"].replace(0, pd.NA)
    ).fillna(0)
    pivot = pd.concat([pivot, total_row], axis=0)

    ordered_columns = sorted([col for col in pivot.columns if col not in {"Total Sales", "Total Profit", "Profit Margin"}])
    ordered_columns.extend(["Total Sales", "Total Profit", "Profit Margin"])
    return pivot[ordered_columns].reset_index().rename(columns={"index": "Region"})


def create_top_customers(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    customer_summary = (
        df.groupby(["Customer ID", "Customer Name", "Segment"], dropna=False)
        .agg(
            total_revenue=("Sales", "sum"),
            total_profit=("Profit", "sum"),
            order_count=("Order ID", "nunique"),
        )
        .reset_index()
    )
    customer_summary["avg_order_value"] = (
        customer_summary["total_revenue"] / customer_summary["order_count"].replace(0, pd.NA)
    ).fillna(0)
    top_customers = customer_summary.sort_values(
        ["total_revenue", "total_profit"],
        ascending=[False, False],
    ).head(20)

    display_df = top_customers[
        ["Customer ID", "Customer Name", "total_revenue", "total_profit", "order_count", "avg_order_value"]
    ].rename(
        columns={
            "total_revenue": "Total Revenue",
            "total_profit": "Total Profit",
            "order_count": "Order Count",
            "avg_order_value": "Avg Order Value",
        }
    )
    lookup_df = customer_summary[["Customer ID", "Segment"]].drop_duplicates().sort_values("Customer ID")
    return display_df.reset_index(drop=True), lookup_df.reset_index(drop=True)


def create_product_performance(df: pd.DataFrame) -> pd.DataFrame:
    products = (
        df.groupby(["Category", "Sub-Category", "Product Name"], dropna=False)
        .agg(
            total_sales=("Sales", "sum"),
            total_profit=("Profit", "sum"),
            order_count=("Order ID", "nunique"),
        )
        .reset_index()
    )
    products["margin_pct"] = safe_margin(products["total_profit"], products["total_sales"])
    return products.rename(
        columns={
            "total_sales": "Sales",
            "total_profit": "Profit",
            "order_count": "Order Count",
            "margin_pct": "Margin %",
        }
    ).sort_values("Margin %", ascending=False)


def write_dataframe(worksheet, dataframe: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> tuple[int, int]:
    headers = list(dataframe.columns)
    for col_offset, header in enumerate(headers, start=start_col):
        worksheet.cell(row=start_row, column=col_offset, value=header)

    for row_offset, (_, row) in enumerate(dataframe.iterrows(), start=start_row + 1):
        for col_offset, value in enumerate(row, start=start_col):
            worksheet.cell(row=row_offset, column=col_offset, value=value)

    return start_row, start_row + len(dataframe)


def apply_number_formats(worksheet, header_row: int, last_row: int) -> None:
    header_map = {
        worksheet.cell(row=header_row, column=col).value: col
        for col in range(1, worksheet.max_column + 1)
    }

    for header, col_index in header_map.items():
        if header is None:
            continue
        normalized = str(header).lower()
        for row in range(header_row + 1, last_row + 1):
            cell = worksheet.cell(row=row, column=col_index)
            if any(keyword in normalized for keyword in ["sales", "profit", "revenue", "value"]):
                cell.number_format = CURRENCY_FORMAT
            elif "margin" in normalized or "growth" in normalized:
                cell.number_format = PERCENT_FORMAT
            elif "orders" in normalized or "count" in normalized:
                cell.number_format = INTEGER_FORMAT


def build_executive_summary_sheet(worksheet, df: pd.DataFrame) -> None:
    worksheet.title = "Executive Summary"
    worksheet.freeze_panes = "A2"

    total_revenue = df["Sales"].sum()
    total_profit = df["Profit"].sum()
    total_orders = df["Order ID"].nunique()
    avg_order_value = total_revenue / total_orders if total_orders else 0

    worksheet["A1"] = "Executive KPI Summary"
    worksheet["A1"].font = Font(size=14, bold=True)

    kpi_rows = [
        ("Total Revenue", total_revenue),
        ("Total Profit", total_profit),
        ("Total Orders", total_orders),
        ("Avg Order Value", avg_order_value),
    ]

    for row_number, (label, value) in enumerate(kpi_rows, start=3):
        worksheet.cell(row=row_number, column=1, value=label)
        worksheet.cell(row=row_number, column=2, value=value)
        worksheet.cell(row=row_number, column=1).fill = KPI_FILL
        worksheet.cell(row=row_number, column=1).font = BOLD_FONT
        worksheet.cell(row=row_number, column=1).border = THIN_BORDER
        worksheet.cell(row=row_number, column=2).border = THIN_BORDER

    worksheet["B3"].number_format = CURRENCY_FORMAT
    worksheet["B4"].number_format = CURRENCY_FORMAT
    worksheet["B5"].number_format = INTEGER_FORMAT
    worksheet["B6"].number_format = CURRENCY_FORMAT

    yearly = create_yearly_summary(df).rename(
        columns={
            "Year": "Year",
            "total_revenue": "Total Revenue",
            "total_profit": "Total Profit",
            "total_orders": "Total Orders",
            "avg_order_value": "Avg Order Value",
            "profit_margin": "Profit Margin %",
            "yoy_growth": "YoY Growth %",
        }
    )
    header_row, last_row = write_dataframe(worksheet, yearly, start_row=9, start_col=1)
    style_header_row(worksheet, header_row)
    style_data_range(worksheet, header_row + 1, last_row, 1, worksheet.max_column)
    apply_number_formats(worksheet, header_row, last_row)

    yoy_col = None
    for col in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=header_row, column=col).value == "YoY Growth %":
            yoy_col = get_column_letter(col)
            break
    if yoy_col and last_row > header_row:
        worksheet.conditional_formatting.add(
            f"{yoy_col}{header_row + 1}:{yoy_col}{last_row}",
            CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="C6EFCE")),
        )
        worksheet.conditional_formatting.add(
            f"{yoy_col}{header_row + 1}:{yoy_col}{last_row}",
            CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE")),
        )

    autofit_columns(worksheet)


def build_regional_analysis_sheet(worksheet, df: pd.DataFrame) -> None:
    worksheet.title = "Regional Analysis"
    worksheet.freeze_panes = "A2"

    pivot = create_regional_pivot(df)
    header_row, last_row = write_dataframe(worksheet, pivot)
    style_header_row(worksheet, header_row)
    style_data_range(worksheet, header_row + 1, last_row, 1, worksheet.max_column)

    last_col_letter = get_column_letter(worksheet.max_column)
    total_row = last_row
    for col in range(2, worksheet.max_column + 1):
        cell = worksheet.cell(row=total_row, column=col)
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL

    header_to_col = {
        worksheet.cell(row=1, column=col).value: col
        for col in range(1, worksheet.max_column + 1)
    }

    for header, col in header_to_col.items():
        if header == "Region":
            continue
        for row in range(2, last_row + 1):
            cell = worksheet.cell(row=row, column=col)
            if header == "Profit Margin":
                cell.number_format = PERCENT_FORMAT
            else:
                cell.number_format = CURRENCY_FORMAT

    profit_margin_col = get_column_letter(header_to_col["Profit Margin"])
    worksheet.conditional_formatting.add(
        f"{profit_margin_col}2:{profit_margin_col}{last_row}",
        ColorScaleRule(
            start_type="num",
            start_value=-0.2,
            start_color="F8696B",
            mid_type="num",
            mid_value=0,
            mid_color="FFEB84",
            end_type="num",
            end_value=0.3,
            end_color="63BE7B",
        ),
    )

    worksheet.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
    autofit_columns(worksheet)


def build_top_customers_sheet(worksheet, df: pd.DataFrame) -> None:
    worksheet.title = "Top Customers"
    worksheet.freeze_panes = "A2"

    top_customers, lookup_df = create_top_customers(df)
    visible_headers = [
        "Customer ID",
        "Customer Name",
        "Segment Lookup",
        "Total Revenue",
        "Total Profit",
        "Order Count",
        "Avg Order Value",
    ]

    for col, header in enumerate(visible_headers, start=1):
        worksheet.cell(row=1, column=col, value=header)

    for row_index, (_, row) in enumerate(top_customers.iterrows(), start=2):
        worksheet.cell(row=row_index, column=1, value=row["Customer ID"])
        worksheet.cell(row=row_index, column=2, value=row["Customer Name"])
        worksheet.cell(
            row=row_index,
            column=3,
            value=f'=VLOOKUP(A{row_index},$J$2:$K${len(lookup_df) + 1},2,FALSE)',
        )
        worksheet.cell(row=row_index, column=4, value=row["Total Revenue"])
        worksheet.cell(row=row_index, column=5, value=row["Total Profit"])
        worksheet.cell(row=row_index, column=6, value=row["Order Count"])
        worksheet.cell(row=row_index, column=7, value=row["Avg Order Value"])

    for lookup_row, (_, row) in enumerate(lookup_df.iterrows(), start=2):
        worksheet.cell(row=lookup_row, column=10, value=row["Customer ID"])
        worksheet.cell(row=lookup_row, column=11, value=row["Segment"])

    style_header_row(worksheet, 1)
    style_data_range(worksheet, 2, len(top_customers) + 1, 1, 7)

    for row in range(2, len(top_customers) + 2):
        if row % 2 == 0:
            for col in range(1, 8):
                worksheet.cell(row=row, column=col).fill = ALT_ROW_FILL

    for col in [4, 5, 7]:
        for row in range(2, len(top_customers) + 2):
            worksheet.cell(row=row, column=col).number_format = CURRENCY_FORMAT

    for row in range(2, len(top_customers) + 2):
        worksheet.cell(row=row, column=6).number_format = INTEGER_FORMAT

    worksheet.column_dimensions["J"].hidden = True
    worksheet.column_dimensions["K"].hidden = True
    autofit_columns(worksheet)


def build_product_performance_sheet(worksheet, df: pd.DataFrame) -> None:
    worksheet.title = "Product Performance"
    worksheet.freeze_panes = "A2"

    products = create_product_performance(df)
    header_row, last_row = write_dataframe(worksheet, products)
    style_header_row(worksheet, header_row)
    style_data_range(worksheet, header_row + 1, last_row, 1, worksheet.max_column)

    header_map = {
        worksheet.cell(row=1, column=col).value: col
        for col in range(1, worksheet.max_column + 1)
    }
    for header in ["Sales", "Profit"]:
        col = header_map[header]
        for row in range(2, last_row + 1):
            worksheet.cell(row=row, column=col).number_format = CURRENCY_FORMAT

    margin_col = header_map["Margin %"]
    for row in range(2, last_row + 1):
        worksheet.cell(row=row, column=margin_col).number_format = PERCENT_FORMAT

    order_count_col = header_map["Order Count"]
    for row in range(2, last_row + 1):
        worksheet.cell(row=row, column=order_count_col).number_format = INTEGER_FORMAT

    worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{last_row}"
    autofit_columns(worksheet)


def main() -> None:
    input_path = resolve_input_path()
    if not input_path.exists():
        raise FileNotFoundError(
            f"Input CSV not found: {input_path}\n"
            f"Place '{DEFAULT_INPUT_FILE}' in the current directory or pass a file path:\n"
            f"python excel_report.py /path/to/superstore_clean.csv"
        )

    print(f"Loading cleaned dataset from: {input_path}")
    df = pd.read_csv(input_path)
    ensure_required_columns(df)
    df = prepare_dataframe(df)

    workbook = Workbook()
    build_executive_summary_sheet(workbook.active, df)
    build_regional_analysis_sheet(workbook.create_sheet(), df)
    build_top_customers_sheet(workbook.create_sheet(), df)
    build_product_performance_sheet(workbook.create_sheet(), df)

    output_path = input_path.with_name(OUTPUT_FILE)
    workbook.save(output_path)
    print(f"Excel report saved to: {output_path}")


if __name__ == "__main__":
    main()
